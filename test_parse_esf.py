import json
import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(__file__))
from parse_esf import extract_paths, parse_time, build_process_tree, parse_esf_jsonl


# ── Minimal ESF event factory ─────────────────────────────────────────────────

def make_event(event_type, event_key, event_data, pid=100, ppid=1, exe="/bin/test",
               time="2026-03-30T09:39:05.814694527Z", action_type="NOTIFY"):
    return {
        "time": time,
        "mach_time": 123456,
        "global_seq_num": 1,
        "seq_num": 1,
        "event_type": event_type,
        "action_type": action_type,
        "action": {},
        "process": {
            "audit_token": {"pid": pid, "ruid": 501, "rgid": 20, "euid": 501},
            "ppid": ppid,
            "session_id": 1,
            "executable": {"path": exe},
            "signing_id": "com.test",
            "team_id": "",
            "is_platform_binary": False,
            "codesigning_flags": 0,
            "start_time": "2026-03-30T09:39:04.000000Z",
        },
        "event": {event_key: event_data},
    }


def make_fork_event(parent_pid, child_pid, child_exe="/bin/test"):
    return make_event(
        event_type=11, event_key="fork",
        event_data={
            "child": {
                "audit_token": {"pid": child_pid, "ruid": 501, "rgid": 20, "euid": 501},
                "ppid": parent_pid,
                "executable": {"path": child_exe},
                "start_time": "2026-03-30T09:39:06.000000Z",
            }
        },
        pid=parent_pid, exe="/bin/parent",
    )


def make_exec_event(pid, ppid, new_exe, args):
    return make_event(
        event_type=9, event_key="exec",
        event_data={
            "target": {
                "audit_token": {"pid": pid, "ruid": 501, "rgid": 20, "euid": 501},
                "ppid": ppid,
                "executable": {"path": new_exe},
                "start_time": "2026-03-30T09:39:06.500000Z",
            },
            "args": args,
        },
        pid=ppid, exe="/bin/parent",
    )


def write_jsonl(path, events, encoding="utf-8"):
    data = "\n".join(json.dumps(e) for e in events)
    raw = data.encode(encoding)
    if encoding == "utf-16":
        # prepend BOM
        raw = b'\xff\xfe' + data.encode("utf-16-le")
    with open(path, "wb") as f:
        f.write(raw)


# ── Tests ─────────────────────────────────────────────────────────────────────

class TestExtractPaths(unittest.TestCase):

    def test_flat_path(self):
        obj = {"path": "/bin/sh", "other": "value"}
        self.assertEqual(extract_paths(obj), ["/bin/sh"])

    def test_nested_path(self):
        obj = {"level1": {"path": "/usr/bin/python"}}
        self.assertEqual(extract_paths(obj), ["/usr/bin/python"])

    def test_multiple_paths(self):
        obj = {"a": {"path": "/bin/a"}, "b": {"path": "/bin/b"}}
        self.assertCountEqual(extract_paths(obj), ["/bin/a", "/bin/b"])

    def test_list_of_objects(self):
        obj = {"items": [{"path": "/bin/x"}, {"path": "/bin/y"}]}
        self.assertCountEqual(extract_paths(obj), ["/bin/x", "/bin/y"])

    def test_empty_path_ignored(self):
        obj = {"path": ""}
        self.assertEqual(extract_paths(obj), [])

    def test_non_string_path_ignored(self):
        obj = {"path": 123}
        self.assertEqual(extract_paths(obj), [])

    def test_empty_dict(self):
        self.assertEqual(extract_paths({}), [])


class TestParseTime(unittest.TestCase):

    def test_nanosecond_timestamp(self):
        utc, local = parse_time("2026-03-30T09:39:05.814694527Z")
        self.assertEqual(utc, "2026-03-30 09:39:05.814694 UTC")
        self.assertIn("2026-03-30", local)

    def test_microsecond_timestamp(self):
        utc, local = parse_time("2026-03-30T09:39:05.814694Z")
        self.assertEqual(utc, "2026-03-30 09:39:05.814694 UTC")

    def test_no_fractional_seconds(self):
        utc, local = parse_time("2026-03-30T09:39:05Z")
        self.assertEqual(utc, "2026-03-30 09:39:05.000000 UTC")

    def test_invalid_returns_empty(self):
        utc, local = parse_time("not-a-time")
        self.assertEqual(utc, "")
        self.assertEqual(local, "")

    def test_empty_string_returns_empty(self):
        utc, local = parse_time("")
        self.assertEqual(utc, "")
        self.assertEqual(local, "")

    def test_none_returns_empty(self):
        utc, local = parse_time(None)
        self.assertEqual(utc, "")
        self.assertEqual(local, "")

    def test_utc_suffix_stripped(self):
        utc, _ = parse_time("2026-01-15T12:00:00.000000Z")
        self.assertTrue(utc.endswith("UTC"))

    def test_non_utc_offset_converted_correctly(self):
        """Timestamp with +05:30 offset must be converted to UTC, not relabelled."""
        utc, _ = parse_time("2026-03-30T09:39:05+05:30")
        self.assertEqual(utc, "2026-03-30 04:09:05.000000 UTC")

    def test_explicit_utc_offset(self):
        """Timestamp with +00:00 explicit UTC offset should round-trip correctly."""
        utc, _ = parse_time("2026-03-30T09:39:05+00:00")
        self.assertEqual(utc, "2026-03-30 09:39:05.000000 UTC")

    def test_nanoseconds_with_non_utc_offset(self):
        """Nanosecond timestamp with non-UTC offset must not lose the offset during truncation."""
        utc, _ = parse_time("2026-03-30T09:39:05.814694527+05:30")
        self.assertEqual(utc, "2026-03-30 04:09:05.814694 UTC")

    def test_nanoseconds_with_negative_offset(self):
        """Nanosecond timestamp with negative offset must convert correctly."""
        utc, _ = parse_time("2026-03-30T04:09:05.814694527-05:30")
        self.assertEqual(utc, "2026-03-30 09:39:05.814694 UTC")


class TestBuildProcessTree(unittest.TestCase):

    def test_fork_creates_child(self):
        events = [make_fork_event(parent_pid=100, child_pid=200, child_exe="/bin/sh")]
        tree = build_process_tree(events)
        pids = [r["pid"] for r in tree]
        self.assertIn(100, pids)
        self.assertIn(200, pids)

    def test_parent_child_relationship(self):
        events = [make_fork_event(parent_pid=100, child_pid=200)]
        tree = build_process_tree(events)
        child = next(r for r in tree if r["pid"] == 200)
        self.assertEqual(child["ppid"], 100)

    def test_exec_updates_exe_and_cmdline(self):
        events = [
            make_fork_event(parent_pid=100, child_pid=200, child_exe="/bin/sh"),
            make_exec_event(pid=200, ppid=100, new_exe="/usr/bin/python3", args=["python3", "script.py"]),
        ]
        tree = build_process_tree(events)
        child = next(r for r in tree if r["pid"] == 200)
        self.assertEqual(child["exe"], "/usr/bin/python3")
        self.assertEqual(child["cmdline"], "python3 script.py")

    def test_tree_depth(self):
        # root(100) -> child(200) -> grandchild(300)
        events = [
            make_fork_event(parent_pid=100, child_pid=200),
            make_fork_event(parent_pid=200, child_pid=300),
        ]
        tree = build_process_tree(events)
        depth = {r["pid"]: r["depth"] for r in tree}
        self.assertEqual(depth[100], 0)
        self.assertEqual(depth[200], 1)
        self.assertEqual(depth[300], 2)

    def test_dfs_order(self):
        # root(100) -> child(200), child(201)
        # child(200) -> grandchild(300)
        events = [
            make_fork_event(parent_pid=100, child_pid=200),
            make_fork_event(parent_pid=100, child_pid=201),
            make_fork_event(parent_pid=200, child_pid=300),
        ]
        tree = build_process_tree(events)
        pids = [r["pid"] for r in tree]
        # 300 should appear before 201 (DFS: go deep before visiting sibling)
        self.assertLess(pids.index(300), pids.index(201))

    def test_empty_events(self):
        tree = build_process_tree([])
        self.assertEqual(tree, [])

    def test_exec_without_prior_fork_creates_entry(self):
        """exec for a pid that was never forked should still appear in tree."""
        # Use ppid=2 so outer process (pid=2, ppid=1 default) becomes a proper root,
        # and the exec target (pid=500, ppid=2) becomes its child.
        events = [make_exec_event(pid=500, ppid=2, new_exe="/usr/bin/curl", args=["curl", "http://x"])]
        tree = build_process_tree(events)
        pids = [r["pid"] for r in tree]
        self.assertIn(500, pids)

    def test_fork_with_null_child_pid_is_ignored(self):
        """fork event where child audit_token has no pid should not crash."""
        event = make_fork_event(parent_pid=100, child_pid=200)
        # Remove pid from child audit_token to simulate missing pid
        event["event"]["fork"]["child"]["audit_token"].pop("pid")
        tree = build_process_tree([event])
        pids = [r["pid"] for r in tree]
        self.assertNotIn(None, pids)

    def test_null_event_field_does_not_crash(self):
        """event dict where 'event' value is JSON null should not raise."""
        event = make_event(9, "exec", {}, pid=100)
        event["event"] = None  # simulate JSON null
        # build_process_tree must not crash
        tree = build_process_tree([event])
        self.assertIsInstance(tree, list)

    def test_sibling_order_in_dfs(self):
        """Children of the same parent should appear in ascending pid order."""
        events = [
            make_fork_event(parent_pid=100, child_pid=201),
            make_fork_event(parent_pid=100, child_pid=200),
        ]
        tree = build_process_tree(events)
        children = [r["pid"] for r in tree if r["ppid"] == 100]
        self.assertEqual(children, sorted(children))

    def test_upsert_exec_updates_only_nonempty_fields(self):
        """exec with empty cmdline should not wipe out a previous cmdline."""
        # Use parent_pid=2 so pid=2 becomes a proper root (its outer process has ppid=1).
        events = [
            make_fork_event(parent_pid=2, child_pid=50, child_exe="/bin/sh"),
            make_exec_event(pid=50, ppid=2, new_exe="/bin/bash", args=["bash", "-c", "ls"]),
            # Second exec with no args should not clear cmdline
            make_exec_event(pid=50, ppid=2, new_exe="/bin/bash", args=[]),
        ]
        tree = build_process_tree(events)
        proc = next(r for r in tree if r["pid"] == 50)
        self.assertEqual(proc["cmdline"], "bash -c ls")

    def test_upsert_none_ppid_does_not_overwrite_valid_ppid(self):
        """A subsequent event with ppid=None (JSON null) must not overwrite a valid integer ppid."""
        # First event establishes pid=10 with ppid=1
        e1 = make_event(43, "lookup", {"lookup": {}}, pid=10, ppid=1)
        # Second event has explicit null ppid (e.g. a malformed or edge-case ESF record)
        e2 = make_event(43, "lookup", {"lookup": {}}, pid=10, ppid=1)
        e2["process"]["ppid"] = None
        tree = build_process_tree([e1, e2])
        proc = next(r for r in tree if r["pid"] == 10)
        self.assertEqual(proc["ppid"], 1)

    def test_upsert_none_start_time_does_not_overwrite_valid_start_time(self):
        """A subsequent event with start_time=None must not overwrite a valid start_time."""
        e1 = make_event(43, "lookup", {"lookup": {}}, pid=10, ppid=1)
        e1["process"]["start_time"] = "2026-03-30T09:00:00.000000Z"
        e2 = make_event(43, "lookup", {"lookup": {}}, pid=10, ppid=1)
        e2["process"]["start_time"] = None
        tree = build_process_tree([e1, e2])
        proc = next(r for r in tree if r["pid"] == 10)
        self.assertEqual(proc["start_time"], "2026-03-30T09:00:00.000000Z")

    def test_self_loop_pid_equals_ppid(self):
        """A process whose ppid equals its own pid should not cause infinite DFS."""
        event = make_event(43, "lookup", {"lookup": {}}, pid=999, ppid=999)
        tree = build_process_tree([event])
        # Should not infinite-loop; pid=999 ppid=999 is in all_pids so not a root → dropped
        pids = [r["pid"] for r in tree]
        # It may or may not appear, but it must terminate
        self.assertIsInstance(tree, list)

    def test_cycle_two_nodes_does_not_hang(self):
        """Two processes forming a ppid cycle should not cause infinite DFS."""
        e1 = make_event(43, "lookup", {"lookup": {}}, pid=10, ppid=20)
        e2 = make_event(43, "lookup", {"lookup": {}}, pid=20, ppid=10)
        tree = build_process_tree([e1, e2])
        # Both ppids are in all_pids, so neither is a root → both dropped safely
        self.assertIsInstance(tree, list)

    def test_null_process_field_does_not_crash(self):
        """event with 'process': null should not crash build_process_tree."""
        event = {"process": None, "event": {"lookup": {}}}
        tree = build_process_tree([event])
        self.assertIsInstance(tree, list)

    def test_null_child_in_fork_does_not_crash(self):
        """fork event where 'child' is null should not crash."""
        event = make_fork_event(parent_pid=100, child_pid=200)
        event["event"]["fork"]["child"] = None
        tree = build_process_tree([event])
        self.assertIsInstance(tree, list)

    def test_null_target_in_exec_does_not_crash(self):
        """exec event where 'target' is null should not crash."""
        event = make_exec_event(pid=200, ppid=100, new_exe="/bin/sh", args=["sh"])
        event["event"]["exec"]["target"] = None
        tree = build_process_tree([event])
        self.assertIsInstance(tree, list)

    def test_null_executable_in_process_does_not_crash(self):
        """process with 'executable': null should not crash."""
        event = make_event(43, "lookup", {"lookup": {}}, pid=100)
        event["process"]["executable"] = None
        tree = build_process_tree([event])
        self.assertIsInstance(tree, list)

class TestParseEsfJsonl(unittest.TestCase):

    def _run_tree_test(self, events, encoding="utf-8"):
        """Helper used by process-tree specific tests."""
        return self._run(events, encoding)

    def _run(self, events, encoding="utf-8"):
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as f:
            write_jsonl(f.name, events, encoding)
            in_path = f.name
        out_path = in_path.replace(".json", ".xlsx")
        try:
            parse_esf_jsonl(in_path, out_path)
            return out_path
        finally:
            os.unlink(in_path)

    def test_basic_output_file_created(self):
        events = [make_event(43, "lookup", {"lookup": {}}, pid=100)]
        out = self._run(events)
        self.assertTrue(os.path.exists(out))
        os.unlink(out)

    def test_event_count_in_sheet(self):
        import openpyxl
        events = [make_event(43, "lookup", {"lookup": {}}, pid=i) for i in range(1, 6)]
        out = self._run(events)
        wb = openpyxl.load_workbook(out)
        data_rows = wb["Events"].max_row - 1  # subtract header
        self.assertEqual(data_rows, 5)
        os.unlink(out)

    def test_both_sheets_present(self):
        import openpyxl
        events = [make_event(43, "lookup", {"lookup": {}}, pid=100)]
        out = self._run(events)
        wb = openpyxl.load_workbook(out)
        self.assertIn("Events", wb.sheetnames)
        self.assertIn("Process Tree", wb.sheetnames)
        os.unlink(out)

    def test_date_time_columns_populated(self):
        import openpyxl
        events = [make_event(43, "lookup", {"lookup": {}}, pid=100,
                             time="2026-03-30T09:39:05.814694527Z")]
        out = self._run(events)
        wb = openpyxl.load_workbook(out)
        ws = wb["Events"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        utc_col = headers.index("date_time_utc") + 1
        val = ws.cell(2, utc_col).value
        self.assertIn("2026-03-30", val)
        self.assertIn("UTC", val)
        os.unlink(out)

    def test_utf16_encoding_supported(self):
        import openpyxl
        events = [make_event(43, "lookup", {"lookup": {}}, pid=100)]
        out = self._run(events, encoding="utf-16")
        wb = openpyxl.load_workbook(out)
        self.assertEqual(wb["Events"].max_row - 1, 1)
        os.unlink(out)

    def test_truncated_file_handled_gracefully(self):
        import openpyxl
        events = [make_event(43, "lookup", {"lookup": {}}, pid=i) for i in range(1, 4)]
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False, mode="wb") as f:
            full = ("\n".join(json.dumps(e) for e in events)).encode("utf-8")
            f.write(full[:-50])  # truncate last 50 bytes
            in_path = f.name
        out_path = in_path.replace(".json", ".xlsx")
        try:
            parse_esf_jsonl(in_path, out_path)
            wb = openpyxl.load_workbook(out_path)
            # Should have parsed at least 2 complete events
            self.assertGreaterEqual(wb["Events"].max_row - 1, 2)
        finally:
            os.unlink(in_path)
            if os.path.exists(out_path):
                os.unlink(out_path)

    def test_fork_exec_appear_in_process_tree(self):
        import openpyxl
        events = [
            make_fork_event(parent_pid=100, child_pid=200, child_exe="/bin/sh"),
            make_exec_event(pid=200, ppid=100, new_exe="/usr/bin/python3", args=["python3"]),
        ]
        out = self._run(events)
        wb = openpyxl.load_workbook(out)
        wt = wb["Process Tree"]
        headers = [wt.cell(1, c).value for c in range(1, wt.max_column + 1)]
        pid_col = headers.index("PID") + 1
        pids = [int(wt.cell(r, pid_col).value) for r in range(2, wt.max_row + 1)]
        self.assertIn(100, pids)
        self.assertIn(200, pids)
        os.unlink(out)

    def test_null_event_field_does_not_crash_jsonl(self):
        """An event with JSON null as the 'event' value must not crash parse_esf_jsonl."""
        import openpyxl
        good_event = make_event(43, "lookup", {"lookup": {}}, pid=100)
        null_event = make_event(9, "exec", {}, pid=200)
        null_event["event"] = None  # simulate "event": null in JSON
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False, mode="w") as f:
            f.write(json.dumps(good_event) + "\n" + json.dumps(null_event))
            in_path = f.name
        out_path = in_path.replace(".json", ".xlsx")
        try:
            parse_esf_jsonl(in_path, out_path)
            wb = openpyxl.load_workbook(out_path)
            # Both events should be written as rows
            self.assertEqual(wb["Events"].max_row - 1, 2)
        finally:
            os.unlink(in_path)
            if os.path.exists(out_path):
                os.unlink(out_path)

    def test_process_tree_indents_children(self):
        """Executable column for child rows should have indentation prefix."""
        import openpyxl
        events = [make_fork_event(parent_pid=100, child_pid=200, child_exe="/bin/child")]
        out = self._run(events)
        wb = openpyxl.load_workbook(out)
        wt = wb["Process Tree"]
        headers = [wt.cell(1, c).value for c in range(1, wt.max_column + 1)]
        exe_col = headers.index("Executable") + 1
        exe_vals = [wt.cell(r, exe_col).value for r in range(2, wt.max_row + 1)]
        child_exe = next(v for v in exe_vals if v and "child" in v)
        self.assertIn("\u2514\u2500", child_exe)  # └─
        os.unlink(out)

    def test_malformed_event_no_process_key(self):
        """Event missing 'process' and 'event' keys should produce a row with empty fields."""
        import openpyxl
        malformed = {"result": {"result_type": 1, "result": {"flags": 0}}}
        good = make_event(43, "lookup", {"lookup": {}}, pid=100)
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False, mode="w") as f:
            f.write(json.dumps(good) + "\n" + json.dumps(malformed))
            in_path = f.name
        out_path = in_path.replace(".json", ".xlsx")
        try:
            parse_esf_jsonl(in_path, out_path)
            wb = openpyxl.load_workbook(out_path)
            self.assertEqual(wb["Events"].max_row - 1, 2)
        finally:
            os.unlink(in_path)
            if os.path.exists(out_path):
                os.unlink(out_path)

    def test_null_process_and_action_in_jsonl(self):
        """Event with 'process': null and 'action': null should not crash parse_esf_jsonl."""
        import openpyxl
        event = {"process": None, "action": None, "event": {"lookup": {}},
                 "time": "2026-03-30T09:39:05.000000Z", "event_type": 43}
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False, mode="w") as f:
            f.write(json.dumps(event))
            in_path = f.name
        out_path = in_path.replace(".json", ".xlsx")
        try:
            parse_esf_jsonl(in_path, out_path)
            wb = openpyxl.load_workbook(out_path)
            self.assertEqual(wb["Events"].max_row - 1, 1)
        finally:
            os.unlink(in_path)
            if os.path.exists(out_path):
                os.unlink(out_path)

    def test_null_executable_in_jsonl(self):
        """Event with 'executable': null inside process should not crash."""
        import openpyxl
        event = make_event(43, "lookup", {"lookup": {}}, pid=100)
        event["process"]["executable"] = None
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False, mode="w") as f:
            f.write(json.dumps(event))
            in_path = f.name
        out_path = in_path.replace(".json", ".xlsx")
        try:
            parse_esf_jsonl(in_path, out_path)
            wb = openpyxl.load_workbook(out_path)
            self.assertEqual(wb["Events"].max_row - 1, 1)
        finally:
            os.unlink(in_path)
            if os.path.exists(out_path):
                os.unlink(out_path)

    def test_no_events_prints_message(self):
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False, mode="w") as f:
            f.write("   ")
            in_path = f.name
        out_path = in_path.replace(".json", ".xlsx")
        try:
            import io
            from contextlib import redirect_stdout
            buf = io.StringIO()
            with redirect_stdout(buf):
                parse_esf_jsonl(in_path, out_path)
            self.assertIn("No events", buf.getvalue())
        finally:
            os.unlink(in_path)
            if os.path.exists(out_path):
                os.unlink(out_path)


if __name__ == "__main__":
    unittest.main(verbosity=2)
